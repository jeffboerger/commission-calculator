"""
Commission Calculator — ERS Platform
=====================================
Scrapes transaction data from ourers.com and calculates commission.

Commission formula:
    damageWaiver = (paid - paidTax - travelFee - surfaceFee - parkFee) / 1.1 * 0.1
    subtotal     = paid - paidTax - travelFee - surfaceFee - parkFee - damageWaiver
    commRev      = subtotal + parkFee + surfaceFee
    commission   = commRev * 0.10

Commission period is based on EVENT DATE, not payment date.
Only transactions where your username appears in the logs as "new order" are included.

Usage:
    Single mode:  python commission_calculator.py --single
    Batch mode:   python commission_calculator.py --batch transactions.txt
    Month mode:   python commission_calculator.py --month "March 2025"
"""

import os
import re
import csv
import argparse
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from datetime import datetime, date
from typing import Optional, Tuple
import calendar
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

load_dotenv()

# ── Config ────────────────────────────────────────────────────────────────────

BASE_URL           = os.getenv("BASE_URL",   "https://yourplatform.ourers.com")
LOGIN_URL          = os.getenv("LOGIN_URL",  "https://yourplatform.ourers.com/cp/login/")
LIST_URL           = os.getenv("LIST_URL",   "https://yourplatform.ourers.com/cp/events/")
USERNAME           = os.getenv("SITE_USERNAME", "")
PASSWORD           = os.getenv("SITE_PASSWORD", "")
MY_USERNAME        = os.getenv("MY_USERNAME", "your.username")
SALES_REPS         = [r.strip() for r in os.getenv("SALES_REPS", MY_USERNAME).split(",")]
EXEMPT_ACTIVATORS  = [r.strip().lower() for r in os.getenv("EXEMPT_ACTIVATORS", "Online Customer,autopay,Admin").split(",")]

COMMISSION_RATE = 0.10


# ── URL helpers ───────────────────────────────────────────────────────────────

def pay_url(order_id):
    return f"{BASE_URL}/cp/review_events/{order_id}/pay/"

def logs_url(order_id):
    return f"{BASE_URL}/cp/review_events/{order_id}/view-logs/"


# ── Auth ──────────────────────────────────────────────────────────────────────

def create_session() -> requests.Session:
    """Log in and return an authenticated session."""
    session = requests.Session()

    payload = {
        "login_username": USERNAME,
        "login_password": PASSWORD,
    }

    resp = session.post(LOGIN_URL, data=payload)
    resp.raise_for_status()
    return session


# ── List page ─────────────────────────────────────────────────────────────────

def fetch_order_ids_for_month(session: requests.Session, year: int, month: int) -> list[str]:
    """
    Fetch the event list page filtered to a month and extract order IDs.

    OPTION A (current): URL date params — update param names if needed.
    OPTION B: If the calendar is JS-driven and the URL doesn't change,
              swap this out for the Playwright stub at the bottom of the file.
    """
    start_date = date(year, month, 1)
    end_date   = date(year, month, calendar.monthrange(year, month)[1])

    # TODO: Confirm exact param names by filtering on the list page and checking the URL
    params = {
        "start_date": start_date.strftime("%Y-%m-%d"),
        "end_date":   end_date.strftime("%Y-%m-%d"),
    }

    resp = session.get(LIST_URL, params=params)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    # TODO: Update selector to match the list page links to individual orders
    # Pattern: links to /cp/review_events/XXXX/
    order_ids = []
    for link in soup.find_all("a", href=True):
        match = re.search(r"/cp/review_events/(\d+)/", link["href"])
        if match:
            oid = match.group(1)
            if oid not in order_ids:
                order_ids.append(oid)

    return order_ids


# ── Pay page scraping ─────────────────────────────────────────────────────────

def scrape_pay_page(session: requests.Session, order_id: str) -> dict:
    """
    Scrape the pay page for dollar amounts and event date.
    Row IDs confirmed from real page inspection.
    """
    resp = session.get(pay_url(order_id))
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    def row_amount(row_id: str, col_index: int = 2) -> float:
        """Extract dollar amount from a table row by its id attribute."""
        row = soup.find("tr", {"id": row_id})
        if not row:
            return 0.0
        cells = row.find_all("td")
        if len(cells) <= col_index:
            return 0.0
        return parse_currency(cells[col_index].text)

    # Dollar fields - confirmed row IDs and cell indices from page inspection
    paid        = row_amount("ers_payment_line_total", col_index=1)       # 2-cell row (colspan)
    paid_tax    = row_amount("ers_line_item_tax", col_index=1)             # cell[1] = tax amount
    subtotal_raw = row_amount("ers_line_item_subtotal", col_index=2)       # read directly from page

    # Optional fees/adjustments - returns 0.0 if row not present on this order
    travel_fee  = row_amount("ers_line_item_travel_fee", col_index=1)
    surface_fee = row_amount("ers_line_item_option_surface_fee", col_index=1)
    park_fee    = row_amount("ers_line_item_option_park-business-non-residential", col_index=1)
    misc_fee    = row_amount("ers_line_item_miscellaneous_fees", col_index=1)
    coupon      = row_amount("ers_line_item_coupon", col_index=1)           # already negative
    discount    = row_amount("ers_line_item_general_discount", col_index=1) # already negative

    # Check if fully paid - Due must be $0.00
    due = row_amount("ers_payment_line_due", col_index=1)
    # Event date - inside .booking-info .delivery-date span
    event_date = None
    date_span  = soup.select_one(".booking-info .delivery-date span")
    if date_span:
        event_date = parse_event_date(date_span.text.strip())

    return {
        "paid":         paid,
        "paid_tax":     paid_tax,
        "subtotal_raw": subtotal_raw,
        "travel_fee":   travel_fee,
        "surface_fee":  surface_fee,
        "park_fee":     park_fee,
        "misc_fee":     misc_fee,
        "coupon":       coupon,
        "discount":     discount,
        "event_date":   event_date,
        "due":          due,
        "not_paid":     due > 0.01,
    }


def parse_event_date(text: str) -> Optional[date]:
    """
    Parse event date from two formats seen in the wild:
      New orders: 'Sat, Mar 14 8:30 -> 10:30 pm'  -> extracted as 'Mar 14 <current year>'
      Old orders: 'Sat, 12/6/2025 11:00 am -> 3:00 pm' -> extracted as '12/6/2025'
    """
    # Try numeric format first: MM/DD/YYYY or M/D/YYYY
    match = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", text)
    if match:
        try:
            return datetime.strptime(match.group(0), "%m/%d/%Y").date()
        except ValueError:
            pass

    # Try month-name format: Mar 14
    match = re.search(r"([A-Za-z]{3,})\s+(\d{1,2})", text)
    if match:
        try:
            month_str = match.group(1)
            day       = int(match.group(2))
            year      = datetime.now().year
            return datetime.strptime(f"{month_str} {day} {year}", "%b %d %Y").date()
        except ValueError:
            pass

    return None


# ── Logs page scraping ────────────────────────────────────────────────────────

def scrape_log_reps(session: requests.Session, order_id: str) -> dict:
    """
    Scrape the logs page and determine which sales reps touched the order.

    Rules:
    - Collect all unique usernames from the logs
    - Filter out exempt accounts (Online Customer, autopay, Admin, etc.)
    - Of the remaining known sales reps:
        1 rep  -> 100% commission
        2 reps -> 50/50 split
    - If MY_USERNAME is not in the list -> not our order

    Returns:
      - reps:    list of known sales rep usernames who touched the order
      - creator: earliest non-exempt username (for display)
    """
    resp = session.get(logs_url(order_id))
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    rows = []
    for row in soup.find_all("tr"):
        cells = row.find_all("td")
        if len(cells) >= 4:
            username = cells[1].text.strip().lstrip()
            action   = cells[3].text.strip()
            if username and username.lower() not in ("account", ""):
                rows.append({"username": username, "action": action})

    if not rows:
        return {"reps": [], "creator": None}

    # Logs are newest-first, reverse to get chronological order
    rows_asc = list(reversed(rows))

    # Find earliest non-exempt username as creator
    creator = None
    for row in rows_asc:
        if row["username"].lower() not in EXEMPT_ACTIVATORS:
            creator = row["username"]
            break

    # Collect unique known sales reps who touched the order (preserve order of appearance)
    seen = []
    for row in rows_asc:
        u = row["username"]
        if u not in seen and u.lower() not in EXEMPT_ACTIVATORS:
            seen.append(u)

    # Filter to only known sales reps
    reps = [u for u in seen if u.lower() in [r.lower() for r in SALES_REPS]]

    return {"reps": reps, "creator": creator}


# ── Commission math ───────────────────────────────────────────────────────────

def calculate_commission(subtotal_raw: float, park_fee: float,
                         surface_fee: float, misc_fee: float) -> dict:
    """
    Commission formula based on page subtotal (rental items only, before all fees):
        damageWaiver = subtotal * 0.10
        commRev      = subtotal + parkFee + surfaceFee + miscFee
        commission   = commRev * 0.10
    """
    damage_waiver = round(subtotal_raw * 0.10, 2)
    comm_rev      = round(subtotal_raw + park_fee + surface_fee + misc_fee, 2)
    commission    = round(comm_rev * COMMISSION_RATE, 2)

    return {
        "damage_waiver": damage_waiver,
        "subtotal":      subtotal_raw,
        "comm_rev":      comm_rev,
        "commission":    commission,
    }


# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_currency(value: str) -> float:
    """Strip currency symbols, handle negatives, and convert to float."""
    cleaned = value.replace("$", "").replace(",", "").replace("−", "-").strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


# ── Full transaction processing ───────────────────────────────────────────────

def process_order(session: requests.Session, order_id: str,
                  month_filter: Optional[Tuple[int, int]] = None) -> Optional[dict]:
    """
    Full pipeline for one order:
      1. Scrape logs → verify username
      2. Scrape pay page → get amounts + event date
      3. Filter by event date if month_filter set
      4. Calculate commission
    """
    # Step 1: scrape logs to find which sales reps touched this order
    log_data = scrape_log_reps(session, order_id)
    reps     = log_data["reps"]
    creator  = log_data["creator"]

    if not reps and creator is None:
        print(f"  ✗  {order_id} - could not determine any reps from logs")
        return None

    # Check if MY_USERNAME is involved
    if not any(MY_USERNAME.lower() in r.lower() for r in reps):
        print(f"  ↷  {order_id} - not yours (reps involved: {creator or 'unknown'})")
        return None

    # Determine split based on how many known sales reps touched it
    other_reps = [r for r in reps if MY_USERNAME.lower() not in r.lower()]
    if len(other_reps) == 0:
        share_pct  = 1.0
        split_flag = ""
    else:
        share_pct  = 0.5
        split_flag = f"50/50 split with {other_reps[0]}"

    # Use creator for display (may be non-rep like Admin)
    activator = reps[0] if reps else creator

    # Step 2: pay page
    pay_data = scrape_pay_page(session, order_id)

    if pay_data.get("not_paid"):
        print(f"  ⏳  {order_id} - not fully paid (due: ${pay_data['due']:,.2f}), including as UNPAID")

    # Step 3: event date filter
    event_date = pay_data["event_date"]
    if event_date is None:
        print(f"  ⚠  {order_id} — could not parse event date, including anyway")
    elif month_filter:
        year, month = month_filter
        if event_date.year != year or event_date.month != month:
            print(f"  ↷  {order_id} — event date {event_date} outside target month")
            return None

    # Step 4: commission math
    calc = calculate_commission(
        pay_data["subtotal_raw"], pay_data["park_fee"],
        pay_data["surface_fee"], pay_data["misc_fee"]
    )

    your_commission = round(calc["commission"] * share_pct, 2)

    return {
        "order_id":         order_id,
        "username":         MY_USERNAME,
        "creator":          creator or "",
        "other_reps":       ", ".join(other_reps) if other_reps else "",
        "share_pct":        f"{int(share_pct * 100)}%",
        "split_flag":       split_flag,
        "event_date":       event_date.strftime("%Y-%m-%d") if event_date else "unknown",
        "paid":             pay_data["paid"],
        "paid_tax":         pay_data["paid_tax"],
        "travel_fee":       pay_data["travel_fee"],
        "surface_fee":      pay_data["surface_fee"],
        "park_fee":         pay_data["park_fee"],
        "misc_fee":         pay_data["misc_fee"],
        "coupon":           pay_data["coupon"],
        "discount":         pay_data["discount"],
        "damage_waiver":    calc["damage_waiver"],
        "subtotal":         calc["subtotal"],
        "comm_rev":         calc["comm_rev"],
        "full_commission":  calc["commission"],
        "your_commission":  your_commission,
        "pay_url":          pay_url(order_id),
        "status":           "UNPAID - Due: $" + f"{pay_data.get('due', 0):,.2f}" if pay_data.get("not_paid") else "PAID",
        "due":              pay_data.get("due", 0.0),
    }


# ── Output ────────────────────────────────────────────────────────────────────

def print_result(r: dict):
    print(f"\n  Order       : {r['order_id']}")
    print(f"  Event Date  : {r['event_date']}")
    print(f"  Creator     : {r['creator']}")
    print(f"  Other Reps  : {r['other_reps'] or 'none'}")
    print(f"  Your Share  : {r['share_pct']}")
    if r['split_flag']:
        print(f"  *** FLAG    : {r['split_flag']}")
    print(f"  Paid        : ${r['paid']:>8.2f}")
    print(f"  Tax         : ${r['paid_tax']:>8.2f}")
    print(f"  Travel Fee  : ${r['travel_fee']:>8.2f}")
    print(f"  Surface Fee : ${r['surface_fee']:>8.2f}")
    print(f"  Park Fee    : ${r['park_fee']:>8.2f}")
    print(f"  Misc Fee    : ${r['misc_fee']:>8.2f}")
    print(f"  Coupon      : ${r['coupon']:>8.2f}")
    print(f"  Discount    : ${r['discount']:>8.2f}")
    print(f"  Dmg Waiver  : ${r['damage_waiver']:>8.2f}")
    print(f"  Subtotal    : ${r['subtotal']:>8.2f}")
    print(f"  Comm Rev    : ${r['comm_rev']:>8.2f}")
    print(f"  Full Comm   : ${r['full_commission']:>8.2f}")
    print(f"  Your Comm   : ${r['your_commission']:>8.2f}")


def save_to_csv(results: list[dict], output_file: str, label: str = ""):
    if not results:
        print("No results to save.")
        return

    total_your_commission = sum(r.get("your_commission", 0) for r in results
                                if r.get("status", "PAID") == "PAID")

    # Main visible columns matching commission report format
    main_fields = [
        "order_id", "username", "status", "paid", "paid_tax",
        "travel_fee", "surface_fee", "park_fee", "misc_fee", "coupon", "discount",
        "damage_waiver", "subtotal", "comm_rev", "your_commission",
        "share_pct", "split_flag", "other_reps",
    ]

    # Hidden columns kept to the right for reference
    hidden_fields = ["full_commission"]

    all_fields = main_fields + hidden_fields

    # Build flat rows with renamed keys to match your spreadsheet headers
    header_map = {
        "order_id":       "orderID",
        "username":       "username",
        "status":         "status",
        "paid":           "paid",
        "paid_tax":       "paidTax",
        "travel_fee":     "travelFee",
        "surface_fee":    "surfaceFee",
        "park_fee":       "parkFee",
        "misc_fee":       "miscFee",
        "coupon":         "coupon",
        "discount":       "discount",
        "damage_waiver":  "damageWaiver",
        "subtotal":       "subtotal",
        "comm_rev":       "commRev",
        "your_commission":"Commission",
        "share_pct":      "sharePct",
        "split_flag":     "splitFlag",
        "other_reps":     "otherReps",
        "full_commission":"fullCommission",
    }

    def fmt(val):
        """Format floats as currency strings."""
        if isinstance(val, float):
            return f"${val:,.2f}"
        return val

    with open(output_file, "w", newline="") as f:
        writer = csv.writer(f)

        period = get_period_label(results, fallback=label)

        # Title row
        writer.writerow([f"{period} Commission", MY_USERNAME])
        writer.writerow([])  # blank row

        # Header row
        writer.writerow([header_map[k] for k in all_fields])

        # Data rows
        for r in results:
            writer.writerow([fmt(r.get(k, "")) for k in all_fields])

        # Blank row then total
        writer.writerow([])
        # Pad to Commission column (index 13) then write total
        padding = [""] * 14
        writer.writerow(padding + [fmt(total_your_commission)])

    print(f"\n  Saved to: {output_file}")


def get_period_label(results: list[dict], fallback: str = "") -> str:
    dates = []
    for r in results:
        ed = r.get("event_date", "")
        if ed and ed != "unknown":
            try:
                dates.append(datetime.strptime(ed, "%Y-%m-%d"))
            except ValueError:
                pass
    if dates:
        return min(dates).strftime("%B %Y")
    return fallback


def save_to_xlsx(results: list[dict], output_file: str, label: str = ""):
    if not results:
        print("No results to save.")
        return

    total_your_commission = sum(r.get("your_commission", 0) for r in results
                                if r.get("status", "PAID") == "PAID")

    period = get_period_label(results, fallback=label)

    main_fields = [
        "order_id", "username", "status", "paid", "paid_tax",
        "travel_fee", "surface_fee", "park_fee", "misc_fee", "coupon", "discount",
        "damage_waiver", "subtotal", "comm_rev", "your_commission",
        "share_pct", "split_flag", "other_reps",
    ]
    header_map = {
        "order_id":        "orderID",
        "username":        "username",
        "status":          "status",
        "paid":            "paid",
        "paid_tax":        "paidTax",
        "travel_fee":      "travelFee",
        "surface_fee":     "surfaceFee",
        "park_fee":        "parkFee",
        "misc_fee":        "miscFee",
        "coupon":          "coupon",
        "discount":        "discount",
        "damage_waiver":   "damageWaiver",
        "subtotal":        "subtotal",
        "comm_rev":        "commRev",
        "your_commission": "Commission",
        "share_pct":       "sharePct",
        "split_flag":      "splitFlag",
        "other_reps":      "otherReps",
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commission"

    # Styles
    title_font    = Font(bold=True, size=13)
    header_font   = Font(bold=True)
    red_font      = Font(color="CC0000")
    red_fill      = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    currency_fmt  = "#,##0.00"
    header_fill   = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font_w = Font(bold=True, color="FFFFFF")

    # Title row
    ws.append([f"{period} Commission", MY_USERNAME])
    ws["A1"].font = title_font
    ws.append([])  # blank

    # Header row
    headers = [header_map[f] for f in main_fields]
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font   = header_font_w
        cell.fill   = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    currency_cols = {header_map[f] for f in main_fields if f in (
        "paid", "paid_tax", "travel_fee", "surface_fee", "park_fee",
        "misc_fee", "coupon", "discount", "damage_waiver", "subtotal",
        "comm_rev", "your_commission"
    )}
    col_letters = {header_map[f]: i+1 for i, f in enumerate(main_fields)}

    for r in results:
        is_unpaid = r.get("status", "PAID") != "PAID"
        row_data  = []
        for f in main_fields:
            val = r.get(f, "")
            if isinstance(val, float) and f == "your_commission" and is_unpaid:
                val = ""  # no commission for unpaid
            row_data.append(val)
        ws.append(row_data)

        data_row = ws.max_row
        for i, f in enumerate(main_fields):
            cell = ws.cell(row=data_row, column=i+1)
            if is_unpaid:
                cell.font = red_font
                cell.fill = red_fill
            if header_map[f] in currency_cols and isinstance(cell.value, float):
                cell.number_format = f'_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

    # Blank row then total
    ws.append([])
    total_row = [""] * 14 + [total_your_commission]
    ws.append(total_row)
    total_cell = ws.cell(row=ws.max_row, column=15)
    total_cell.font   = Font(bold=True)
    total_cell.number_format = f'_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    wb.save(output_file)
    print(f"\n  Saved to: {output_file}")


def save_to_pdf(results: list[dict], output_file: str, label: str = ""):
    if not results:
        print("No results to save.")
        return

    total_your_commission = sum(r.get("your_commission", 0) for r in results
                                if r.get("status", "PAID") == "PAID")
    period = get_period_label(results, fallback=label)

    # Clean columns for PDF (drop split/flag columns)
    pdf_fields = [
        "order_id", "username", "status", "paid", "paid_tax",
        "travel_fee", "surface_fee", "park_fee", "misc_fee", "coupon", "discount",
        "damage_waiver", "subtotal", "comm_rev", "your_commission",
    ]
    header_map = {
        "order_id":        "Order ID",
        "username":        "Rep",
        "status":          "Status",
        "paid":            "Paid",
        "paid_tax":        "Tax",
        "travel_fee":      "Travel",
        "surface_fee":     "Surface",
        "park_fee":        "Park",
        "misc_fee":        "Misc",
        "coupon":          "Coupon",
        "discount":        "Discount",
        "damage_waiver":   "Dmg Waiver",
        "subtotal":        "Subtotal",
        "comm_rev":        "Comm Rev",
        "your_commission": "Commission",
    }

    def fmt(val, field):
        if field == "your_commission" and isinstance(val, float):
            return ""  # filled below per row
        if isinstance(val, float):
            return f"${val:,.2f}"
        return str(val) if val else ""

    doc = SimpleDocTemplate(
        output_file,
        pagesize=landscape(letter),
        leftMargin=0.5*inch, rightMargin=0.5*inch,
        topMargin=0.5*inch, bottomMargin=0.5*inch
    )

    styles    = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Normal"],
                                 fontSize=13, fontName="Helvetica-Bold", spaceAfter=4)
    sub_style   = ParagraphStyle("sub", parent=styles["Normal"],
                                 fontSize=9, textColor=colors.HexColor("#888888"), spaceAfter=12)

    story = []
    story.append(Paragraph(f"{period} Commission", title_style))
    story.append(Paragraph(f"{MY_USERNAME}  |  Generated {datetime.now().strftime('%m/%d/%Y %I:%M %p')}", sub_style))

    # Build table data
    header_row = [header_map[f] for f in pdf_fields]
    table_data  = [header_row]

    for r in results:
        is_unpaid = r.get("status", "PAID") != "PAID"
        row = []
        for f in pdf_fields:
            val = r.get(f, "")
            if f == "your_commission":
                row.append("" if is_unpaid else (f"${val:,.2f}" if isinstance(val, float) else str(val)))
            elif isinstance(val, float):
                row.append(f"${val:,.2f}")
            else:
                row.append(str(val) if val else "")
        table_data.append(row)

    # Total row
    padding = [""] * (len(pdf_fields) - 1)
    table_data.append(padding + [f"${total_your_commission:,.2f}"])

    # Auto-size columns to fit landscape letter (10 inches usable)
    n_cols   = len(pdf_fields)
    col_w    = 10.0 * inch / n_cols
    col_widths = [col_w] * n_cols

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Find unpaid row indices (offset by 1 for header)
    unpaid_rows = [i+1 for i, r in enumerate(results) if r.get("status", "PAID") != "PAID"]
    total_row_idx = len(results) + 1

    style_cmds = [
        # Header
        ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#1F3864")),
        ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, -1), 7),
        ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUND",(0, 1), (-1, -2), [colors.white, colors.HexColor("#F5F8FF")]),
        ("GRID",         (0, 0), (-1, -2), 0.25, colors.HexColor("#CCCCCC")),
        ("LINEABOVE",    (0, total_row_idx), (-1, total_row_idx), 1, colors.HexColor("#1F3864")),
        ("FONTNAME",     (0, total_row_idx), (-1, total_row_idx), "Helvetica-Bold"),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
    ]

    # Red rows for unpaid
    for row_idx in unpaid_rows:
        style_cmds.append(("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#FFE0E0")))
        style_cmds.append(("TEXTCOLOR",  (0, row_idx), (-1, row_idx), colors.HexColor("#CC0000")))

    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)

    doc.build(story)
    print(f"\n  Saved to: {output_file}")


def print_summary(results: list[dict], label: str = ""):
    total_comm_rev        = sum(r.get("comm_rev", 0) for r in results)
    total_your_commission = sum(r.get("your_commission", 0) for r in results)
    flagged               = [r for r in results if r.get("split_flag")]
    print(f"\n{'─'*45}")
    if label:
        print(f"  {label}")
    print(f"  Rep                    : {MY_USERNAME}")
    print(f"  Transactions processed : {len(results)}")
    print(f"  Total Sales (Comm Rev) : ${total_comm_rev:,.2f}")
    print(f"  Your Total Commission  : ${total_your_commission:,.2f}")
    if flagged:
        print(f"  *** Flagged for review : {len(flagged)} order(s)")
        for r in flagged:
            print(f"      - Order {r['order_id']}: {r['split_flag']}")
    print(f"{'─'*45}")


# ── Modes ─────────────────────────────────────────────────────────────────────

def run_single(session: requests.Session):
    results = []
    print("Enter order IDs one at a time. Type 'q' or 'quit' when done.\n")

    while True:
        order_id = input("Order ID (or q to finish): ").strip()
        if order_id.lower() in ("q", "quit", ""):
            break

        result = process_order(session, order_id)
        if result:
            print_result(result)
            results.append(result)

    if not results:
        print("No results to save.")
        return

    print_summary(results, label="Manual Entry")
    print("\nSave as: (1) Excel  (2) PDF  (3) CSV  (4) All  (n) Don't save")
    save = input("Choice: ").strip().lower()
    if save in ("1", "2", "3", "4", "excel", "pdf", "csv", "all"):
        period = get_period_label(results, fallback="commission")
        ts     = datetime.now().strftime("%Y%m%d_%H%M")
        base   = f"jeff_commission_{period.replace(' ', '_')}_{ts}"
        if save in ("1", "excel", "4", "all"):
            save_to_xlsx(results, f"{base}.xlsx", label="Manual Entry")
        if save in ("2", "pdf", "4", "all"):
            save_to_pdf(results, f"{base}.pdf", label="Manual Entry")
        if save in ("3", "csv", "4", "all"):
            save_to_csv(results, f"{base}.csv", label="Manual Entry")


def run_batch(session: requests.Session, input_file: str):
    with open(input_file, "r") as f:
        order_ids = [line.strip() for line in f if line.strip()]

    print(f"Processing {len(order_ids)} orders...\n")
    results = []
    for oid in order_ids:
        r = process_order(session, oid)
        if r:
            print_result(r)
            results.append(r)

    output_file = input_file.replace(".txt", "_results.csv").replace(".csv", "_results.csv")
    save_to_csv(results, output_file, label=f"Batch: {input_file}")
    print_summary(results, label=f"Batch: {input_file}")


def run_month(session: requests.Session, month_str: str):
    for fmt in ("%B %Y", "%b %Y", "%m/%Y", "%m-%Y"):
        try:
            dt = datetime.strptime(month_str.strip(), fmt)
            break
        except ValueError:
            continue
    else:
        print(f"Could not parse month: '{month_str}'. Try 'March 2025' or '03/2025'.")
        return

    year, month = dt.year, dt.month
    month_label = dt.strftime("%B %Y")

    print(f"Fetching orders for {month_label} (by event date)...\n")
    order_ids = fetch_order_ids_for_month(session, year, month)

    if not order_ids:
        print("No orders found for that month.")
        return

    print(f"Found {len(order_ids)} orders. Processing...\n")
    results = []
    for oid in order_ids:
        r = process_order(session, oid, month_filter=(year, month))
        if r:
            print_result(r)
            results.append(r)

    output_file = f"commission_{dt.strftime('%Y_%m')}.csv"
    save_to_csv(results, output_file, label=month_label)
    print_summary(results, label=f"Commission for {month_label}")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Commission Calculator — ERS Platform")
    parser.add_argument("--batch", metavar="FILE",       help="Process a file of order IDs")
    parser.add_argument("--month", metavar="MONTH_YEAR", help="Pull all orders by event month, e.g. 'March 2025'")
    args = parser.parse_args()

    print("Logging in...")
    session = create_session()
    print("Logged in.\n")

    if args.batch:
        run_batch(session, args.batch)
    elif args.month:
        run_month(session, args.month)
    else:
        run_single(session)


if __name__ == "__main__":
    main()


# ── Playwright stub (if list page calendar is JS-driven) ──────────────────────
#
# pip install playwright && playwright install chromium
#
# from playwright.sync_api import sync_playwright
#
# def fetch_order_ids_for_month_playwright(year: int, month: int) -> list[str]:
#     with sync_playwright() as p:
#         browser = p.chromium.launch(headless=True)
#         page    = browser.new_page()
#         page.goto(LOGIN_URL)
#         page.fill("input[name='username']", USERNAME)
#         page.fill("input[name='password']", PASSWORD)
#         page.click("button[type='submit']")
#         page.wait_for_load_state("networkidle")
#         page.goto(LIST_URL)
#         # TODO: interact with calendar widget to set month filter
#         html    = page.content()
#         browser.close()
#     soup = BeautifulSoup(html, "html.parser")
#     order_ids = []
#     for link in soup.find_all("a", href=True):
#         match = re.search(r"/cp/review_events/(\d+)/", link["href"])
#         if match:
#             oid = match.group(1)
#             if oid not in order_ids:
#                 order_ids.append(oid)
#     return order_ids
